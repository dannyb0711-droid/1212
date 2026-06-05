#!/usr/bin/env python3
"""
inject_monthly_actuals.py
─────────────────────────────────────────────────────────────────────
동화약품 OTC 대시보드 — 월간 실적 주입 도구
─────────────────────────────────────────────────────────────────────

사용법:
  python inject_monthly_actuals.py <일일실적보고.xlsx> [data.js]

기능:
  일일실적보고 Excel의 'DB' 시트에서 당월 실적(영업부/팀/코스별)을
  읽어 data.js의 _meta.currentMonthActuals에 주입합니다.
  이를 통해 대시보드 당월 달성률이 일일실적보고와 정확히 일치합니다.

왜 필요한가?
  - data.js의 당월 실적(ms)은 기본적으로 dailySales '정가' 기준으로 집계
  - 일일실적보고는 SAP DB 기준 '실적금액' 사용 → 두 기준이 다름
  - currentMonthActuals를 주입하면 대시보드가 이 값을 우선 사용

매월 업데이트 절차:
  1. DW OTC 데이터 변환기로 data.js 생성
  2. 일일실적보고 Excel(당월) 준비
  3. 이 스크립트 실행:
       python inject_monthly_actuals.py 일일실적보고_26_06_.xlsx data.js
  4. 수정된 data.js를 GitHub에 push → Netlify 자동 배포
"""

import sys
import json
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl이 설치되어 있지 않습니다.")
    print("  pip install openpyxl")
    sys.exit(1)


def extract_actuals_from_excel(excel_path: str) -> dict:
    """
    일일실적보고 Excel의 'DB' 시트에서 당월 실적을 추출합니다.

    Returns:
        {
          'byDept':   { '111': 633884354, ... },
          'byTeam':   { '1111': 105034097, ... },
          'byCourse': { '111101': 24954576, ... }
        }
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if 'DB' not in wb.sheetnames:
        raise ValueError(
            f"'DB' 시트를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}"
        )

    ws = wb['DB']
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("'DB' 시트가 비어 있습니다.")

    # 헤더 확인 (행1): 영업부(0), 영업지점(1), 영업코스(2), ..., 실적금액(7)
    header = rows[0]
    expected = ('영업부', '영업지점', '영업코스')
    for i, col in enumerate(expected):
        if str(header[i] or '').strip() != col:
            raise ValueError(
                f"DB 시트 헤더 불일치. {i}번째 열이 '{col}'이어야 하는데 '{header[i]}'입니다."
            )

    dept_actuals   = defaultdict(float)
    team_actuals   = defaultdict(float)
    course_actuals = defaultdict(float)

    skipped = 0
    for row_num, row in enumerate(rows[1:], start=2):
        if not row[0]:
            continue

        dept   = str(row[0]).strip()
        team   = str(row[1]).strip()
        course = str(row[2]).strip()
        actual = row[7]  # 실적금액 열

        if actual is None:
            skipped += 1
            continue
        if not isinstance(actual, (int, float)):
            skipped += 1
            continue

        dept_actuals[dept]     += actual
        team_actuals[team]     += actual
        course_actuals[course] += actual

    if skipped > 0:
        print(f"  [참고] 실적금액이 비어있는 행 {skipped}건 건너뜀")

    return {
        'byDept':   {k: int(v) for k, v in sorted(dept_actuals.items())},
        'byTeam':   {k: int(v) for k, v in sorted(team_actuals.items())},
        'byCourse': {k: int(v) for k, v in sorted(course_actuals.items())},
    }


def inject_into_data_js(data_js_path: str, cma: dict) -> str:
    """
    data.js의 _meta.currentMonthActuals:null 을 실제 값으로 교체합니다.

    Returns:
        수정된 파일 내용(str)
    """
    with open(data_js_path, 'r', encoding='utf-8') as f:
        content = f.read()

    old_str = '"currentMonthActuals":null'
    if old_str not in content:
        # 이미 값이 있을 수도 있음 — 기존 값을 교체
        pattern = r'"currentMonthActuals"\s*:\s*\{[^}]*(?:\{[^}]*\}[^}]*)?\}'
        if re.search(pattern, content):
            new_cma_str = '"currentMonthActuals":' + json.dumps(cma, ensure_ascii=False)
            content = re.sub(pattern, new_cma_str, content, count=1)
            print("  [참고] 기존 currentMonthActuals 값을 새 값으로 교체했습니다.")
            return content
        else:
            raise ValueError(
                "'currentMonthActuals' 필드를 data.js에서 찾을 수 없습니다.\n"
                "DW OTC 데이터 변환기 v2.0으로 생성된 data.js인지 확인하세요."
            )

    new_cma_str = '"currentMonthActuals":' + json.dumps(cma, ensure_ascii=False)
    return content.replace(old_str, new_cma_str, 1)


def validate_result(cma: dict) -> None:
    """주입 결과 요약을 출력합니다."""
    by_dept = cma.get('byDept', {})
    total_actual = sum(by_dept.values())

    dept_names = {
        '111': '약국1부', '112': '약국2부', '113': '약국3부',
        '118': 'C/G_OTC', '121': '도매부'
    }

    print("\n  ┌─────────────────────────────────────────────────┐")
    print("  │  주입된 currentMonthActuals (byDept) 요약       │")
    print("  ├──────────┬──────────┬──────────────────────────┤")
    print("  │ 영업부   │ 부서명   │ 당월 실적 (원)           │")
    print("  ├──────────┼──────────┼──────────────────────────┤")
    for dc in sorted(by_dept.keys()):
        name = dept_names.get(dc, '?')
        val  = by_dept[dc]
        print(f"  │ {dc:<8} │ {name:<8} │ {val:>24,} │")
    print("  ├──────────┴──────────┼──────────────────────────┤")
    print(f"  │ 전체 합계            │ {total_actual:>24,} │")
    print("  └─────────────────────┴──────────────────────────┘")

    team_cnt   = len(cma.get('byTeam', {}))
    course_cnt = len(cma.get('byCourse', {}))
    print(f"\n  팀 수: {team_cnt}개, 코스 수: {course_cnt}개")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    excel_path   = sys.argv[1]
    data_js_path = sys.argv[2] if len(sys.argv) > 2 else 'data.js'

    # ── 경로 확인
    if not Path(excel_path).exists():
        print(f"ERROR: Excel 파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)
    if not Path(data_js_path).exists():
        print(f"ERROR: data.js 파일을 찾을 수 없습니다: {data_js_path}")
        sys.exit(1)

    print(f"\n[1/3] Excel에서 당월 실적 추출 중...")
    print(f"      파일: {excel_path}")
    cma = extract_actuals_from_excel(excel_path)

    print(f"\n[2/3] data.js에 currentMonthActuals 주입 중...")
    print(f"      파일: {data_js_path}")
    new_content = inject_into_data_js(data_js_path, cma)

    # 원본 백업
    backup_path = data_js_path + '.bak'
    with open(backup_path, 'w', encoding='utf-8') as f:
        with open(data_js_path, 'r', encoding='utf-8') as orig:
            f.write(orig.read())
    print(f"      백업 생성: {backup_path}")

    # 저장
    with open(data_js_path, 'w', encoding='utf-8') as f:
        f.write(new_content)
    print(f"      저장 완료: {data_js_path}")

    print(f"\n[3/3] 결과 검증:")
    validate_result(cma)

    print("\n✅ 완료! 이제 data.js를 GitHub에 push하면 대시보드가 업데이트됩니다.")
    print("   달성률이 일일실적보고와 정확히 일치합니다.\n")


if __name__ == '__main__':
    main()
